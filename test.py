import os
import pandas as pd
from tqdm import tqdm
import utils
import Constant
import xlwt
import openpyxl

rp = Constant.resultPath
ap = Constant.addressPath
tx_path = Constant.tornadoTxPath
outer_tx_path = Constant.addressOuterTxPath


def calDelta(txDf, value):
    delta = []
    length = txDf.shape[0]
    last_tx = dict()
    ntxTimeMap = dict()
    num = 0
    for idx in tqdm(range(length)):
        tx = txDf.iloc[idx].to_dict()
        if not last_tx:
            last_tx = tx
            continue
        if tx['type'] == last_tx['type']:
            num += 1
        else:
            if ntxTimeMap.get(num):
                if (ntxTimeMap[num]['type'] == 'Deposit' and last_tx['type'] == 'Withdrawal' and ntxTimeMap[num]['from'] != last_tx['to']) or (ntxTimeMap[num]['type'] == 'Withdrawal' and last_tx['type'] == 'Deposit' and ntxTimeMap[num]['to'] != last_tx['from']):
                    delta.append({"time":(last_tx['timeStamp']-ntxTimeMap[num]['timeStamp'])//60,'n':num})
            ntxTimeMap[num] = last_tx
            last_tx = dict()
            num = 0
    save_path = tx_path + '统计数据/deltaAnalysis0201.xlsx'
    index = 0
    if os.path.exists(save_path):
        wb = openpyxl.load_workbook(save_path)
        index = len(wb.sheetnames)
    else:
        wb = openpyxl.Workbook()
    if value in wb.sheetnames:
        sheet = wb[value]
    else:
        sheet = wb.create_sheet(title=value, index=index)
        wb.active = index
    sheet.title = value
    label = ['time','n']
    for id,l in enumerate(label):
        sheet.cell(1,id+1,l)
    for i in range(len(delta)):
        sheet.cell(i+2, 1, delta[i][label[0]])
        sheet.cell(i+2, 2, delta[i][label[1]])
    wb.save(save_path)


def calTime(txDf, value):
    dwTime = []
    wdTime = []
    ddTime = []
    wwTime = []
    last_tx = dict()
    length = txDf.shape[0]
    for idx in tqdm(range(length)):
        tx = txDf.iloc[idx].to_dict()
        if not last_tx:
            last_tx = tx
            continue
        if tx['type'] == 'Deposit':
            if last_tx['type'] == 'Deposit' and last_tx['from']!=tx['from']:
                ddTime.append((tx['timeStamp'] - last_tx['timeStamp'])//60)
            elif last_tx['type'] == 'Withdrawal' and last_tx['from']!=tx['to']:
                wdTime.append((tx['timeStamp'] - last_tx['timeStamp'])//60)
        elif tx['type'] == 'Withdrawal':
            if last_tx['type'] == 'Deposit' and last_tx['to']!=tx['from']:
                dwTime.append((tx['timeStamp'] - last_tx['timeStamp'])//60)
            elif last_tx['type'] == 'Withdrawal' and last_tx['to']!=tx['to']:
                wwTime.append((tx['timeStamp'] - last_tx['timeStamp'])//60)
        last_tx = tx
    save_path = tx_path + '统计数据/timeAnalysis0201.xlsx'
    index = 0
    if os.path.exists(save_path):
        wb = openpyxl.load_workbook(save_path)
        index = len(wb.sheetnames)
    else:
        wb = openpyxl.Workbook()
    if value in wb.sheetnames:
        sheet = wb[value]
    else:
        sheet = wb.create_sheet(title=value, index=index)
        wb.active = index
    sheet.title = value
    label = ['dwTime', 'wdTime', 'ddTime', 'wwTime']
    times = [dwTime, wdTime, ddTime, wwTime]
    for id, item in enumerate(label):
        sheet.cell(1, id+1, item)
        for i in range(len(times[id])):
            sheet.cell(i+2, id+1, times[id][i])
    wb.save(save_path)


def analysisTime(tp):
    for file_dir in os.listdir(tp + "CSV"):
        value = file_dir.split('_')[0]
        df = utils.readDataFromCsv(tp + 'csv/' + file_dir)
        calTime(df, value)
        calDelta(df, value)


if __name__ == '__main__':
    # analysisTime(tx_path)
    ts = 1576523323000
    import datetime
    ts = ts*1.0
    print(type(ts))
    dt = datetime.datetime.utcfromtimestamp(ts)